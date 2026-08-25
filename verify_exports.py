from pathlib import Path
from openpyxl import load_workbook

root = Path('/home/ubuntu/Downloads')
xlsx_files = sorted(root.glob('recoverai-transactions-*.xlsx'), key=lambda path: path.stat().st_mtime, reverse=True)
if xlsx_files:
    path = xlsx_files[0]
    workbook = load_workbook(path, read_only=True, data_only=True)
    print({'xlsx': path.name, 'sheets': workbook.sheetnames, 'rows': workbook['Transactions'].max_row, 'columns': workbook['Transactions'].max_column})
else:
    print({'xlsx': 'not found'})

pdf_files = sorted(root.glob('recoverai-analytics-report-*.pdf'), key=lambda path: path.stat().st_mtime, reverse=True)
print({'pdf': pdf_files[0].name if pdf_files else 'not found'})
